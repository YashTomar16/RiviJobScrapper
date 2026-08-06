from __future__ import annotations

import csv
import json
import re
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlparse

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from rivi.models import Company as CompanyRow

REQUIRED_CSV_FIELDS = ("company_name", "category", "website", "career_page", "career_page_status")


@dataclass
class CompanyRecord:
    company_name: str
    category: str
    website: str
    career_page: str = ""
    career_page_status: str = ""
    career_page_source: str = ""
    skip: bool = False
    skip_reason: str = ""

    @property
    def is_eligible(self) -> bool:
        return bool(self.career_page.strip()) and not self.skip


@dataclass
class ImportResult:
    source: str
    total: int
    inserted: int
    updated: int
    with_website: int
    with_career_page: int
    missing_website: int
    missing_career_page: int
    by_category: dict[str, int]
    companies: list[CompanyRecord]


def normalize_website(raw: str | None) -> str:
    if not raw:
        return ""
    w = str(raw).strip()
    if not w or w.lower() in {"n/a", "na", "-", "none"}:
        return ""
    if not w.startswith(("http://", "https://")):
        w = "https://" + w
    return w.rstrip("/")


def normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip())


def load_companies_csv(path: Path) -> list[CompanyRecord]:
    if not path.exists():
        raise FileNotFoundError(f"Companies CSV not found: {path}")

    companies: list[CompanyRecord] = []
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise ValueError(f"CSV has no header row: {path}")

        missing_cols = [c for c in REQUIRED_CSV_FIELDS if c not in reader.fieldnames]
        if missing_cols:
            raise ValueError(
                f"CSV missing required columns {missing_cols}; found {list(reader.fieldnames)}"
            )

        for i, row in enumerate(reader, start=2):
            name = normalize_name(row.get("company_name") or "")
            if not name:
                raise ValueError(f"Row {i}: company_name is required")
            companies.append(
                CompanyRecord(
                    company_name=name,
                    category=(row.get("category") or "").strip(),
                    website=normalize_website(row.get("website")),
                    career_page=(row.get("career_page") or "").strip(),
                    career_page_status=(row.get("career_page_status") or "").strip(),
                    career_page_source=(row.get("career_page_source") or "").strip(),
                    skip=(row.get("skip") or "").strip().lower() in {"1", "true", "yes", "y"},
                    skip_reason=(row.get("skip_reason") or "").strip(),
                )
            )
    return companies


def load_companies_excel(path: Path) -> list[CompanyRecord]:
    if not path.exists():
        raise FileNotFoundError(f"Excel not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    companies: list[CompanyRecord] = []
    seen: set[str] = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [str(h).strip() if h is not None else "" for h in next(rows)]
        except StopIteration:
            continue

        name_idx = headers.index("Investors") if "Investors" in headers else 0
        website_idx = headers.index("Website") if "Website" in headers else None

        for row in rows:
            if not row or name_idx >= len(row) or not row[name_idx]:
                continue
            name = normalize_name(str(row[name_idx]))
            if name in seen:
                continue
            seen.add(name)
            website_raw = row[website_idx] if website_idx is not None and website_idx < len(row) else None
            companies.append(
                CompanyRecord(
                    company_name=name,
                    category=sheet_name,
                    website=normalize_website(website_raw),
                )
            )

    return companies


def upsert_companies(session: Session, records: list[CompanyRecord]) -> tuple[int, int]:
    inserted = 0
    updated = 0
    now = datetime.now(timezone.utc)

    for rec in records:
        # Match on name + category so the same company can live in multiple cohorts.
        existing = session.scalar(
            select(CompanyRow).where(
                CompanyRow.name == rec.company_name,
                CompanyRow.category == rec.category,
            )
        )
        if existing is None:
            session.add(
                CompanyRow(
                    name=rec.company_name,
                    category=rec.category,
                    website=rec.website,
                    career_page=rec.career_page,
                    career_page_status=rec.career_page_status,
                    career_page_source=rec.career_page_source
                    or ("auto" if rec.career_page else ""),
                    skip=rec.skip,
                    skip_reason=rec.skip_reason,
                    created_at=now,
                    updated_at=now,
                )
            )
            inserted += 1
            continue

        existing.category = rec.category or existing.category
        existing.website = rec.website or existing.website
        # Preserve manual career pages unless the incoming record is also manual
        if existing.career_page_source == "manual" and rec.career_page_source != "manual":
            pass
        elif rec.career_page:
            existing.career_page = rec.career_page
            existing.career_page_status = rec.career_page_status or existing.career_page_status
            if rec.career_page_source:
                existing.career_page_source = rec.career_page_source
        elif rec.career_page_status and not existing.career_page:
            existing.career_page_status = rec.career_page_status

        if rec.skip:
            existing.skip = True
            existing.skip_reason = rec.skip_reason or existing.skip_reason
        existing.updated_at = now
        updated += 1

    session.commit()
    return inserted, updated


def export_companies_csv(session: Session, path: Path) -> int:
    rows = session.scalars(select(CompanyRow).order_by(CompanyRow.category, CompanyRow.name)).all()
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "company_name",
        "category",
        "website",
        "career_page",
        "career_page_status",
        "career_page_source",
        "skip",
        "skip_reason",
    ]
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in rows:
            writer.writerow(
                {
                    "company_name": r.name,
                    "category": r.category,
                    "website": r.website,
                    "career_page": r.career_page,
                    "career_page_status": r.career_page_status,
                    "career_page_source": r.career_page_source,
                    "skip": "true" if r.skip else "false",
                    "skip_reason": r.skip_reason,
                }
            )
    return len(rows)


def export_companies_json(session: Session, path: Path) -> int:
    rows = session.scalars(select(CompanyRow).order_by(CompanyRow.category, CompanyRow.name)).all()
    payload = [
        {
            "company_name": r.name,
            "category": r.category,
            "website": r.website,
            "career_page": r.career_page,
            "career_page_status": r.career_page_status,
            "career_page_source": r.career_page_source,
            "skip": r.skip,
            "skip_reason": r.skip_reason,
        }
        for r in rows
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return len(payload)


def summarize_records(source: str, records: list[CompanyRecord], inserted: int, updated: int) -> ImportResult:
    by_category: dict[str, int] = {}
    with_website = 0
    with_career = 0
    for c in records:
        by_category[c.category or "unknown"] = by_category.get(c.category or "unknown", 0) + 1
        if c.website:
            with_website += 1
        if c.career_page:
            with_career += 1
    return ImportResult(
        source=source,
        total=len(records),
        inserted=inserted,
        updated=updated,
        with_website=with_website,
        with_career_page=with_career,
        missing_website=len(records) - with_website,
        missing_career_page=len(records) - with_career,
        by_category=by_category,
        companies=records,
    )


def import_from_csv(session: Session, path: Path) -> ImportResult:
    records = load_companies_csv(path)
    inserted, updated = upsert_companies(session, records)
    return summarize_records(str(path.resolve()), records, inserted, updated)


def import_from_excel(session: Session, path: Path) -> ImportResult:
    records = load_companies_excel(path)
    # Excel has no career pages — only upsert name/category/website; preserve existing careers in DB
    inserted, updated = upsert_companies(session, records)
    # Re-read DB counts for accurate career stats after upsert
    rows = session.scalars(select(CompanyRow)).all()
    db_records = [
        CompanyRecord(
            company_name=r.name,
            category=r.category,
            website=r.website,
            career_page=r.career_page,
            career_page_status=r.career_page_status,
            career_page_source=r.career_page_source,
            skip=r.skip,
            skip_reason=r.skip_reason,
        )
        for r in rows
    ]
    return summarize_records(str(path.resolve()), db_records, inserted, updated)


def _resolve_company_row(
    session: Session,
    company_name: str,
    category: str | None = None,
) -> CompanyRow:
    q = select(CompanyRow).where(CompanyRow.name == company_name)
    if category is not None:
        q = q.where(CompanyRow.category == category)
    rows = list(session.scalars(q))
    if not rows:
        hint = f" in category {category!r}" if category else ""
        raise LookupError(f"Company not found: {company_name}{hint}")
    if len(rows) > 1 and category is None:
        cats = ", ".join(sorted({r.category for r in rows}))
        raise LookupError(
            f"Multiple rows for {company_name!r} ({cats}); pass category to disambiguate"
        )
    return rows[0]


def set_career_page_manual(
    session: Session,
    company_name: str,
    career_page: str,
    category: str | None = None,
) -> CompanyRow:
    row = _resolve_company_row(session, company_name, category)
    row.career_page = career_page.strip()
    row.career_page_status = "manual:ok"
    row.career_page_source = "manual"
    row.skip = False
    row.updated_at = datetime.now(timezone.utc)
    session.commit()
    return row


def set_company_skip(
    session: Session,
    company_name: str,
    skip: bool,
    reason: str = "",
    category: str | None = None,
) -> CompanyRow:
    row = _resolve_company_row(session, company_name, category)
    row.skip = skip
    row.skip_reason = reason if skip else ""
    row.updated_at = datetime.now(timezone.utc)
    session.commit()
    return row


def find_company(
    session: Session,
    company_name: str,
    category: str | None = None,
) -> CompanyRow | None:
    try:
        return _resolve_company_row(session, company_name, category)
    except LookupError:
        return None


def list_companies(session: Session) -> list[CompanyRow]:
    return list(session.scalars(select(CompanyRow).order_by(CompanyRow.category, CompanyRow.name)))


def domain_key(website: str) -> str:
    if not website:
        return ""
    host = urlparse(website).netloc.lower()
    return host[4:] if host.startswith("www.") else host


def import_result_to_dict(result: ImportResult, *, include_companies: bool = False) -> dict:
    payload = {
        "source": result.source,
        "total": result.total,
        "inserted": result.inserted,
        "updated": result.updated,
        "with_website": result.with_website,
        "with_career_page": result.with_career_page,
        "missing_website": result.missing_website,
        "missing_career_page": result.missing_career_page,
        "by_category": result.by_category,
        "eligible_for_ingest": result.with_career_page,
    }
    if include_companies:
        payload["companies"] = [asdict(c) for c in result.companies]
    return payload


def write_import_report(result: ImportResult, path: Path, *, include_companies: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(import_result_to_dict(result, include_companies=include_companies), indent=2),
        encoding="utf-8",
    )


# Back-compat alias used by Phase 0 callers
def import_companies(path: Path) -> ImportResult:
    records = load_companies_csv(path)
    return summarize_records(str(path.resolve()), records, inserted=0, updated=0)
